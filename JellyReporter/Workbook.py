from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def create_workbook():
    """Creates and returns a new Excel workbook."""
    return Workbook()

def save_workbook(workbook, output_file):
    """Saves the workbook with the specified filename."""
    workbook.save(output_file)

def apply_styles(workbook, sheet_name):
    """
    Applies styles to every table in the specified sheet.
    Renames specific headers if present, sets all text to Arial 10pt, headers have a black background with white bold text,
    rows alternate between white and grey backgrounds, and column widths are adjusted to fit the longest data or header entry.
    """
    sheet = workbook[sheet_name]

    # Header mapping for renaming
    header_rename_map = {
        "first_name": "First Name",
        "last_name": "Last Name",
        "email": "Email",
        "timestamp": "Timestamp",
        "time": "Time",
        "action": "Action",
        "address": "IP Address",
        "user-agent": "User Agent",
        "message": "Message",
    }

    # Iterate over all tables in the sheet
    for table in sheet.tables.values():
        # Get the table's range coordinates
        start_cell, end_cell = table.ref.split(":")
        start_row = sheet[start_cell].row
        end_row = sheet[end_cell].row
        start_col = sheet[start_cell].column
        end_col = sheet[end_cell].column

        # Apply header styles (first row of the table)
        header_row = sheet[start_row]
        for col_idx in range(start_col, end_col + 1):
            cell = sheet.cell(row=start_row, column=col_idx)
            original_header = cell.value
            # Rename headers if they match the specified keys and are present
            if original_header in header_rename_map:
                cell.value = header_rename_map[original_header]
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="000001", end_color="000001", fill_type="solid")

        # Apply alternating row colors (data rows within the table range)
        for row_idx in range(start_row + 1, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = Font(name="Arial", size=10)
                # Apply alternating fill colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3",
                                            fill_type="solid")  # Light grey for even rows
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                                            fill_type="solid")  # White for odd rows

        # Adjust column widths based on the longest entry in each column (including headers)
        for col_idx in range(start_col, end_col + 1):
            max_length = 0
            # Check header length
            header_cell = sheet.cell(row=start_row, column=col_idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))

            # Check data length in each cell of the column
            for row_idx in range(start_row + 1, end_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Add padding to the width
            adjusted_width = max_length + 2
            column_letter = sheet.cell(row=start_row, column=col_idx).column_letter
            sheet.column_dimensions[column_letter].width = adjusted_width


def apply_conditional_formatting(sheet, dataframe, condition_columns, start_row=2, start_col=4):
    """
    Applies conditional formatting to specific columns based on 'Yes' or 'No' values.

    Parameters:
    sheet: The worksheet to apply formatting.
    dataframe: The pandas DataFrame with the data.
    condition_columns: List of column names in the DataFrame to apply conditional formatting.
    start_row: Row to start applying formatting (defaults to 2 for typical DataFrame header offset).
    start_col: Starting column number for condition columns in the sheet.
    """
    red_fill = PatternFill(start_color='FF1919', end_color='FF1919', fill_type='solid')
    green_fill = PatternFill(start_color='50C878', end_color='50C878', fill_type='solid')

    for index, row in dataframe.iterrows():
        for col_offset, action in enumerate(condition_columns):
            cell = sheet.cell(row=index + start_row, column=start_col + col_offset)
            if row[action] == 'Yes':
                cell.fill = red_fill
            else:
                cell.fill = green_fill

def add_sheet(workbook, sheet_name):
    """Adds a sheet to the workbook with the given name."""
    workbook.create_sheet(title=sheet_name)

def save_dataframe_to_sheet(workbook, sheet_name, table_name, dataframe, columns=None, start_row=1, start_col=1):
    """
    Saves selected columns of the DataFrame to the specified sheet in the workbook at a given start position.

    Parameters:
    workbook: The workbook where the data will be saved.
    sheet_name: Name of the sheet where data will be saved.
    table_name: Name of the table to create.
    dataframe: The pandas DataFrame containing data to save.
    columns: List of column names to save. If None, all columns are saved.
    start_row: Starting row for the table (default is 1).
    start_col: Starting column for the table (default is 1).
    """
    sheet = workbook[sheet_name]

    # Filter the DataFrame to include only the specified columns that are actually present in the DataFrame
    if columns:
        existing_columns = [col for col in columns if col in dataframe.columns]
        data = dataframe[existing_columns]
    else:
        data = dataframe

    # Write headers
    for col_idx, header in enumerate(data.columns, start=start_col):
        sheet.cell(row=start_row, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(data.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row_data, start=start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Define table boundaries
    end_row = start_row + len(data)
    end_col = start_col + len(data.columns) - 1
    table_ref = f"{sheet.cell(row=start_row, column=start_col).coordinate}:{sheet.cell(row=end_row, column=end_col).coordinate}"

    # Create and style the table
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    sheet.add_table(table)
